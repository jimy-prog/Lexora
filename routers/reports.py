from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import date
from io import BytesIO
from database import get_db, Group, Student, Lesson, Attendance
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from finance_rules import get_group_epl

router = APIRouter(prefix="/reports")

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

@router.get("/export/monthly")
def export_monthly(month: str = None, db: Session = Depends(get_db)):
    today = date.today()
    if month:
        y, m = map(int, month.split("-"))
        month_start = date(y, m, 1)
    else:
        month_start = today.replace(day=1)
    me = date(month_start.year, month_start.month + 1, 1) if month_start.month < 12 else date(month_start.year + 1, 1, 1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Report"

    header_fill = PatternFill("solid", start_color="1F3864", fgColor="1F3864")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    calc_fill   = PatternFill("solid", start_color="F2F2F2", fgColor="F2F2F2")

    ws.merge_cells("A1:F1")
    ws["A1"].value = f"Monthly Finance Report — {month_start.strftime('%B %Y')}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    for col, txt in enumerate(["Group","Countable Lessons","Price/Lesson (UZS)","Earn/Lesson (UZS)","Teacher Income (UZS)","Notes"], start=1):
        c = ws.cell(row=3, column=col, value=txt)
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()

    row = 4
    total_income = 0
    groups = db.query(Group).all()
    for g in groups:
        cnt = db.query(Attendance).join(Lesson).filter(
            Lesson.date >= month_start, Lesson.date < me,
            Lesson.status == "Held", Attendance.status == "Present",
            Lesson.group_id == g.id
        ).count()
        epl = get_group_epl(db, g)
        ppl = epl
        inc = cnt * epl
        total_income += inc
        for col, val in enumerate([g.name, cnt, ppl, epl, inc, g.notes or ""], start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name="Calibri", size=10)
            c.fill = calc_fill; c.border = thin_border()
            c.alignment = Alignment(horizontal="center" if col > 1 else "left")
        row += 1

    ws.cell(row=row, column=1, value="TOTAL").font = Font(name="Calibri", bold=True, size=11)
    ws.cell(row=row, column=5, value=total_income).font = Font(name="Calibri", bold=True, size=11, color="276221")
    ws.cell(row=row, column=5).fill = PatternFill("solid", start_color="C6EFCE", fgColor="C6EFCE")

    for col, width in zip("ABCDEF", [20, 18, 20, 20, 22, 30]):
        ws.column_dimensions[col].width = width

    # Attendance sheet
    ws2 = wb.create_sheet("Attendance Detail")
    lessons = db.query(Lesson).filter(Lesson.date >= month_start, Lesson.date < me).order_by(Lesson.date).all()
    students = db.query(Student).filter(Student.active == True).all()
    for col, txt in enumerate(["Date","Group","Student","Attendance","Lesson Status","Countable"], start=1):
        c = ws2.cell(row=1, column=col, value=txt)
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(horizontal="center"); c.border = thin_border()
    r = 2
    for lesson in lessons:
        for att in db.query(Attendance).filter(Attendance.lesson_id == lesson.id).all():
            countable = 1 if att.status == "Present" and lesson.status == "Held" else 0
            for col, val in enumerate([str(lesson.date), lesson.group.name, att.student.name, att.status, lesson.status, countable], start=1):
                c = ws2.cell(row=r, column=col, value=val)
                c.font = Font(name="Calibri", size=10); c.border = thin_border()
                c.alignment = Alignment(horizontal="center" if col != 3 else "left")
            r += 1

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"report_{month_start.strftime('%Y_%m')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})
